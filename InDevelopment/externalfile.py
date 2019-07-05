import pandas as pd
import openpyxl as opxl
from operator import itemgetter

toclst = []
colhead = []
datalst = []

# This gets the text file.
def gettxt():
    df = pd.read_csv('/Users/shared/SampleFiles/metadata2', header=None)
    return df

# This will get the file, initial we have a default file, later we will allow them to input
def getfile():
    #wb = opxl.load_workbook('/Users/xxx/Downloads/Excel Import Text.xlsx')
    wb = opxl.load_workbook('/Users/shared/SampleFiles/Excel Import Text.xlsx')
    #wb = opxl.load_workbook('/Users/shared/SampleFiles/ExcelExport/GTest.xlsx')
    return wb

def getdataheader(df):
    myhead = df[df.iloc[:, 0].str.contains('ead')]
    myhead = myhead.to_string().strip()
    myhead = myhead[20:]
    return myhead

def getdata(df):
    myhead = df[df.iloc[:, 0].str.contains('ata:')]
    myhead = myhead.to_string().strip()
    myhead = myhead[18:]
    return myhead

def returnheader(ws, headerrange):
    for item in headerrange:
        item = item.split('.', 1)[-1][:-1]
        if item.endswith('>'):
           item = item[:-1]
        else:
             item = item
        if len(item) > 1:
             colhead.append(ws[item].value)

    return colhead

def returndata(ws, datarange):
    for item in datarange:
        item = item.split('.', 1)[-1][:-1]

        if item.endswith('>'):
            item = item[:-1]
        elif item.endswith(')'):
            item = item[:-2]
        else:
            item = item
        if len(item) > 1:
            datalst.append(ws[item].value)
    return datalst

def chunks(listy, numy):
    # For item i in a range that is a length of l,
    for i in range(0, len(listy), numy):
        # Create an index range for l of n items:
        yield listy[i:i + numy]

df = gettxt()
print (df)
wb = getfile()

ws = wb['data']

headerrange = getdataheader(df)
headerrange = str(ws[str(headerrange)]).split(",")

#

colprint = returnheader(ws, headerrange)
print('The column headers are:')
print(colprint)


#
#
datarange = getdata(df)
datarange = str(ws[str(datarange)]).split(",")

datalist = returndata(ws, datarange)

# #This chunks up the data so that it prints more elegantly
finalListy = list(chunks(datalist, len(colhead)))

i = 0
while i < len(finalListy):
    print(finalListy[i])
    i += 1

df = pd.DataFrame(finalListy, columns=colprint)

df.to_csv('example.csv', index=False)






